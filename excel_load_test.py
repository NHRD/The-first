import os
import openpyxl

"""
def load_excel(file_name):
    directory = os.getcwd()
    tag_path = directory + "\\" + file_name
    load_excel_path = os.path.join(tag_path)
    return load_excel_path

file_name = "manager_list.xlsx"
"""

wb = openpyxl.load_workbook("manager_list.xlsx")
sheet = wb.get_sheet_by_name("Sheet1")
manager_list = []
for i in range(1,20):
    if sheet.cell(row=i, column=1).value != None:
        manager_list.append(sheet.cell(row=i, column=1).value)
    else:
        break

print(manager_list)